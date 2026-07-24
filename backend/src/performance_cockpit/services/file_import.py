import csv
import io
import re
from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from performance_cockpit.schemas import ImportResult
from performance_cockpit.services.csv_import import import_csv_text

NORMALIZED_HEADERS = {
    "metric_key",
    "metric_name",
    "organizational_unit",
    "period_start",
    "period_end",
    "value",
}
WIDE_METRICS = {
    "vvl": ("vvl", "Vertragsverlängerungen", "Anzahl", "sum"),
    "bnt": ("bnt", "BNT", "Anzahl", "sum"),
    "mobile": ("mobile", "Mobile", "Anzahl", "sum"),
    "vvl mobile": ("vvl_mobile", "VVL Mobile", "Anzahl", "sum"),
    "angebote": ("offers", "Angebote", "Anzahl", "sum"),
    "angebotsquote": ("offer_rate", "Angebotsquote", "Prozent", "average"),
    "calls": ("calls", "Calls", "Anzahl", "sum"),
    "bbcr": ("bbcr", "BBCR", "Prozent", "average"),
    "bewertungen": ("ratings", "Bewertungen", "Anzahl", "sum"),
    "tnps": ("tnps", "TNPS", "Punkte", "average"),
    "cs": ("cs", "CS", "Punkte", "average"),
    "total fix": ("total_fix", "Total Fix", "Prozent", "average"),
    "auflegerquote": ("hangup_rate", "Auflegerquote", "Prozent", "average"),
    "cht": ("cht", "CHT", "Sekunden", "average"),
    "aht": ("aht", "AHT", "Sekunden", "average"),
    "acw": ("acw", "ACW", "Sekunden", "average"),
    "fb quote": ("feedback_rate", "FB Quote", "Prozent", "average"),
}
PERCENT_METRICS = {"angebotsquote", "bbcr", "total fix", "auflegerquote", "fb quote"}
GERMAN_MONTHS = {
    "januar": 1,
    "februar": 2,
    "märz": 3,
    "maerz": 3,
    "april": 4,
    "mai": 5,
    "juni": 6,
    "juli": 7,
    "august": 8,
    "september": 9,
    "oktober": 10,
    "november": 11,
    "dezember": 12,
}


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


def _normalized_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _header_map(worksheet: Worksheet, row_number: int) -> dict[str, int]:
    return {
        _normalized_header(cell.value): cell.column
        for cell in worksheet[row_number]
        if _normalized_header(cell.value)
    }


def _find_wide_header(worksheet: Worksheet) -> int | None:
    for row_number in range(1, min(worksheet.max_row, 100) + 1):
        headers = _header_map(worksheet, row_number)
        if {"teamleiter", "mitarbeiter", "epa"}.issubset(headers) and len(
            WIDE_METRICS.keys() & headers.keys()
        ) >= 3:
            return row_number
    return None


def _organization_name(values: dict[str, object]) -> str:
    employee = str(values.get("mitarbeiter") or "").strip()
    epa = str(values.get("epa") or "").strip()
    team_lead = str(values.get("teamleiter") or "").strip()
    identity = employee or (f"EPA {epa}" if epa else "")
    if team_lead and identity:
        return f"{team_lead} · {identity}"
    return identity or team_lead


def _metric_value(cell: Cell, header: str) -> Decimal | None:
    if cell.data_type == "e" or cell.value in (None, ""):
        return None
    try:
        value = Decimal(str(cell.value).replace(",", "."))
    except (ArithmeticError, ValueError):
        return None
    if header in PERCENT_METRICS and "%" in cell.number_format and abs(value) <= 1:
        value *= 100
    return value


def _wide_rows(
    worksheet: Worksheet,
    header_row: int,
    period: date,
) -> list[dict[str, object]]:
    headers = _header_map(worksheet, header_row)
    rows: list[dict[str, object]] = []
    for row_number in range(header_row + 1, worksheet.max_row + 1):
        identity_values = {
            key: worksheet.cell(row_number, headers[key]).value
            for key in ("teamleiter", "mitarbeiter", "epa")
        }
        organization = _organization_name(identity_values)
        if not organization:
            continue
        for header, definition in WIDE_METRICS.items():
            if header not in headers:
                continue
            value = _metric_value(worksheet.cell(row_number, headers[header]), header)
            if value is None:
                continue
            key, name, unit, aggregation = definition
            rows.append(
                {
                    "metric_key": key,
                    "metric_name": name,
                    "description": f"{name} aus dem Performance-Report",
                    "unit": unit,
                    "aggregation": aggregation,
                    "organizational_unit": organization,
                    "period_start": period.isoformat(),
                    "period_end": period.isoformat(),
                    "value": value,
                    "target_value": "",
                }
            )
    return rows


def _summary_rows(
    worksheet: Worksheet,
    employee_header_row: int,
    period: date,
) -> list[dict[str, object]]:
    for header_row in range(employee_header_row - 1, 0, -1):
        headers = _header_map(worksheet, header_row)
        if "epa" not in headers or len(WIDE_METRICS.keys() & headers.keys()) < 3:
            continue
        value_row = header_row + 1
        organization = str(worksheet.cell(value_row, headers["epa"]).value or "").strip()
        if not organization:
            continue
        rows: list[dict[str, object]] = []
        for header, definition in WIDE_METRICS.items():
            if header not in headers:
                continue
            value = _metric_value(worksheet.cell(value_row, headers[header]), header)
            if value is None:
                continue
            key, name, unit, aggregation = definition
            rows.append(
                {
                    "metric_key": key,
                    "metric_name": name,
                    "description": f"{name} aus der Tageszusammenfassung",
                    "unit": unit,
                    "aggregation": aggregation,
                    "organizational_unit": organization,
                    "period_start": period.isoformat(),
                    "period_end": period.isoformat(),
                    "value": value,
                    "target_value": "",
                }
            )
        return rows
    return []


def _historical_rows(worksheet: Worksheet, organization: str) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for header_row in range(1, min(worksheet.max_row, 20) + 1):
        headers = _header_map(worksheet, header_row)
        if not {"aht", "acw"}.issubset(headers):
            continue
        month_column = min(headers["aht"], headers["acw"]) - 1
        for row_number in range(header_row + 1, worksheet.max_row + 1):
            month_name = _normalized_header(worksheet.cell(row_number, month_column).value)
            month = GERMAN_MONTHS.get(month_name)
            if month is None:
                if rows:
                    break
                continue
            period_start = date(date.today().year, month, 1)
            period_end = date(date.today().year, month, monthrange(date.today().year, month)[1])
            for header in ("aht", "acw"):
                value = _metric_value(worksheet.cell(row_number, headers[header]), header)
                if value is None:
                    continue
                key, name, unit, aggregation = WIDE_METRICS[header]
                rows.append(
                    {
                        "metric_key": key,
                        "metric_name": name,
                        "description": f"{name} aus dem Monatsverlauf",
                        "unit": unit,
                        "aggregation": aggregation,
                        "organizational_unit": organization,
                        "period_start": period_start.isoformat(),
                        "period_end": period_end.isoformat(),
                        "value": value,
                        "target_value": "",
                    }
                )
        break
    return rows


def _wide_workbook_csv(worksheet: Worksheet) -> str | None:
    header_row = _find_wide_header(worksheet)
    if header_row is None:
        return None
    period = date.today()
    rows = _wide_rows(worksheet, header_row, period)
    if not rows:
        rows = _summary_rows(worksheet, header_row, period)
    organization = str(next((row["organizational_unit"] for row in rows), "Gesamt"))
    rows.extend(_historical_rows(worksheet, organization))
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=[
            "metric_key",
            "metric_name",
            "description",
            "unit",
            "aggregation",
            "organizational_unit",
            "period_start",
            "period_end",
            "value",
            "target_value",
        ],
        lineterminator="\n",
    )
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue()


def import_xlsx_bytes(session: Session, file_name: str, content: bytes) -> ImportResult:
    workbook = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    try:
        worksheet = workbook.active
        first_row = {
            _normalized_header(cell.value)
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
            if cell.value is not None
        }
        if not NORMALIZED_HEADERS.issubset(first_row):
            wide_csv = _wide_workbook_csv(worksheet)
            if wide_csv is not None:
                return import_csv_text(session, file_name, wide_csv)
        rows = worksheet.iter_rows(values_only=True)
        output = io.StringIO()
        writer = csv.writer(output, lineterminator="\n")
        for row in rows:
            writer.writerow([_cell_text(value) for value in row])
        # Both file formats intentionally share validation and idempotent persistence.
        return import_csv_text(session, file_name, output.getvalue())
    finally:
        workbook.close()
