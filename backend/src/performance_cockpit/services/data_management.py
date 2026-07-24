import csv
import io
import json
import sqlite3
import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import delete, select
from sqlalchemy.engine import make_url
from sqlalchemy.orm import Session

from performance_cockpit.models import ImportBatch, Measurement, MetricDefinition
from performance_cockpit.schemas import ImportBatchRead, ImportError
from performance_cockpit.services.metrics import get_dashboard

EXPORT_COLUMNS = [
    "metric_key",
    "metric_name",
    "description",
    "unit",
    "aggregation",
    "organizational_unit",
    "period_start",
    "period_end",
    "value",
    "target_value",
    "source",
]
REQUIRED_BACKUP_TABLES = {"metric_definitions", "measurements", "import_batches"}


def sqlite_database_path(database_url: str) -> Path:
    url = make_url(database_url)
    if url.drivername != "sqlite+pysqlite" or not url.database or url.database == ":memory:":
        raise ValueError("Local data management requires a file-based SQLite database")
    return Path(url.database)


def import_history(session: Session, limit: int = 20) -> list[ImportBatchRead]:
    batches = list(
        session.scalars(select(ImportBatch).order_by(ImportBatch.created_at.desc()).limit(limit))
    )
    return [
        ImportBatchRead(
            id=batch.id,
            file_name=batch.file_name,
            status=batch.status,
            total_rows=batch.total_rows,
            imported_rows=batch.imported_rows,
            failed_rows=batch.failed_rows,
            created_at=batch.created_at,
            errors=[
                ImportError.model_validate(error)
                for error in json.loads(batch.error_details or "[]")
            ],
        )
        for batch in batches
    ]


def export_measurements_csv(session: Session) -> str:
    measurements = ordered_measurements(session)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=EXPORT_COLUMNS, lineterminator="\n")
    writer.writeheader()
    for item in measurements:
        writer.writerow(measurement_row(item))
    return output.getvalue()


def ordered_measurements(session: Session) -> list[Measurement]:
    return list(
        session.scalars(
            select(Measurement)
            .join(Measurement.metric)
            .order_by(
                Measurement.metric_key,
                Measurement.organizational_unit,
                Measurement.period_start,
            )
        )
    )


def measurement_row(item: Measurement) -> dict[str, object]:
    return {
        "metric_key": item.metric_key,
        "metric_name": item.metric.display_name,
        "description": item.metric.description,
        "unit": item.metric.unit,
        "aggregation": item.metric.aggregation,
        "organizational_unit": item.organizational_unit,
        "period_start": item.period_start.isoformat(),
        "period_end": item.period_end.isoformat(),
        "value": item.value,
        "target_value": item.target_value if item.target_value is not None else "",
        "source": item.source,
    }


def export_measurements_xlsx(session: Session) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Kennzahlen"
    sheet.append(EXPORT_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="176C53")
    for item in ordered_measurements(session):
        sheet.append([measurement_row(item)[column] for column in EXPORT_COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
        sheet.column_dimensions[column[0].column_letter].width = width
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_text(value: object) -> bytes:
    text = str(value).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    return text.encode("cp1252", errors="replace")


def _simple_pdf(lines: list[tuple[str, int]]) -> bytes:
    commands = [b"BT", b"/F1 20 Tf", b"50 790 Td"]
    for index, (line, size) in enumerate(lines):
        if index:
            commands.append(b"0 -24 Td")
        commands.extend([f"/F1 {size} Tf".encode(), b"(" + _pdf_text(line) + b") Tj"])
    commands.append(b"ET")
    stream = b"\n".join(commands)
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        (
            b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
            b"/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>"
        ),
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >>",
        b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream",
    ]
    document = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for number, content in enumerate(objects, start=1):
        offsets.append(len(document))
        document.extend(f"{number} 0 obj\n".encode() + content + b"\nendobj\n")
    xref = len(document)
    document.extend(f"xref\n0 {len(objects) + 1}\n".encode())
    document.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        document.extend(f"{offset:010d} 00000 n \n".encode())
    document.extend(
        (
            f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n"
        ).encode()
    )
    return bytes(document)


def export_dashboard_pdf(
    session: Session,
    organizational_unit: str,
    date_from: date | None = None,
    date_to: date | None = None,
) -> bytes:
    dashboard = get_dashboard(session, organizational_unit, date_from, date_to)
    period = f"{dashboard.period_start or '-'} bis {dashboard.period_end or '-'}"
    lines = [
        ("Performance Cockpit", 20),
        (f"Bericht: {organizational_unit}", 14),
        (f"Zeitraum: {period}", 10),
        ("", 10),
    ]
    lines.extend(
        (
            f"{summary.display_name}: {summary.value} {summary.unit} "
            f"(Ziel: {summary.target_value if summary.target_value is not None else '-'})",
            11,
        )
        for summary in dashboard.summaries[:25]
    )
    return _simple_pdf(lines)


def create_backup(database_path: Path) -> bytes:
    with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as temporary:
        backup_path = Path(temporary.name)
    try:
        with (
            sqlite3.connect(database_path) as source,
            sqlite3.connect(backup_path) as destination,
        ):
            source.backup(destination)
        return backup_path.read_bytes()
    finally:
        backup_path.unlink(missing_ok=True)


def validate_backup(backup_path: Path) -> None:
    try:
        with sqlite3.connect(backup_path) as connection:
            integrity = connection.execute("PRAGMA integrity_check").fetchone()
            tables = {
                row[0]
                for row in connection.execute(
                    "SELECT name FROM sqlite_master WHERE type = 'table'"
                ).fetchall()
            }
    except sqlite3.DatabaseError as error:
        raise ValueError("Backup is not a valid SQLite database") from error
    if not integrity or integrity[0] != "ok":
        raise ValueError("Backup database failed its integrity check")
    if not REQUIRED_BACKUP_TABLES.issubset(tables):
        raise ValueError("Backup does not contain Performance Cockpit data")


def restore_backup(database_path: Path, content: bytes) -> None:
    with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as temporary:
        temporary.write(content)
        backup_path = Path(temporary.name)
    try:
        validate_backup(backup_path)
        with (
            sqlite3.connect(backup_path) as source,
            sqlite3.connect(database_path) as destination,
        ):
            source.backup(destination)
    finally:
        backup_path.unlink(missing_ok=True)


def reset_data(session: Session) -> None:
    session.execute(delete(Measurement))
    session.execute(delete(MetricDefinition))
    session.execute(delete(ImportBatch))
    session.commit()
