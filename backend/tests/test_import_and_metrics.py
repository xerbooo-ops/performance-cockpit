from decimal import Decimal
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy.orm import Session

from performance_cockpit.models import Measurement

VALID_CSV = "\n".join(
    [
        "metric_key,metric_name,description,unit,aggregation,organizational_unit,"
        "period_start,period_end,value,target_value",
        "handled_cases,Bearbeitete Vorgänge,Abgeschlossene Vorgänge,Anzahl,sum,"
        "Service Nord,2026-07-01,2026-07-07,110,100",
        "handled_cases,Bearbeitete Vorgänge,Abgeschlossene Vorgänge,Anzahl,sum,"
        "Service Nord,2026-07-08,2026-07-14,95,100",
        "quality_rate,Qualitätsquote,Positiv bewertete Vorgänge,Prozent,average,"
        "Service Nord,2026-07-01,2026-07-07,92.5,90",
        "",
    ]
)


def test_csv_import_exposes_metrics_and_summary(client: TestClient) -> None:
    response = client.post(
        "/api/v1/imports/csv",
        files={"file": ("sample.csv", VALID_CSV, "text/csv")},
    )

    assert response.status_code == 200
    assert response.json()["status"] == "completed"
    assert response.json()["imported_rows"] == 3

    metrics = client.get("/api/v1/metrics")
    assert [metric["key"] for metric in metrics.json()] == ["handled_cases", "quality_rate"]

    summary = client.get(
        "/api/v1/metrics/handled_cases/summary",
        params={"organizational_unit": "Service Nord"},
    )
    assert summary.status_code == 200
    assert summary.json() == {
        "metric_key": "handled_cases",
        "display_name": "Bearbeitete Vorgänge",
        "organizational_unit": "Service Nord",
        "period_start": "2026-07-01",
        "period_end": "2026-07-14",
        "unit": "Anzahl",
        "aggregation": "sum",
        "value": "205.00",
        "target_value": "200.00",
        "deviation": "5.00",
        "attainment_percent": "102.50",
        "measurement_count": 2,
    }


def test_reimport_updates_existing_measurement(client: TestClient, session: Session) -> None:
    client.post("/api/v1/imports/csv", files={"file": ("first.csv", VALID_CSV, "text/csv")})
    updated = VALID_CSV.replace(",110,100", ",120,100")

    response = client.post(
        "/api/v1/imports/csv",
        files={"file": ("updated.csv", updated, "text/csv")},
    )

    assert response.status_code == 200
    measurements = session.query(Measurement).all()
    assert len(measurements) == 3
    assert measurements[0].value == Decimal("120")


def test_csv_import_returns_row_level_validation_errors(client: TestClient) -> None:
    invalid_csv = VALID_CSV + (
        "bad key,Ungültig,Fehler,Anzahl,sum,Service Nord,2026-07-14,2026-07-01,10,20\n"
    )

    response = client.post(
        "/api/v1/imports/csv",
        files={"file": ("invalid.csv", invalid_csv, "text/csv")},
    )

    body = response.json()
    assert response.status_code == 200
    assert body["status"] == "completed_with_errors"
    assert body["imported_rows"] == 3
    assert body["failed_rows"] == 1
    assert body["errors"][0]["row"] == 5


def test_csv_import_accepts_empty_optional_target(client: TestClient) -> None:
    csv_without_target = VALID_CSV.replace("92.5,90", "92.5,")

    response = client.post(
        "/api/v1/imports/csv",
        files={"file": ("without-target.csv", csv_without_target, "text/csv")},
    )

    assert response.status_code == 200
    assert response.json()["status"] == "completed"


def test_missing_metric_and_invalid_upload_return_clear_errors(client: TestClient) -> None:
    metric_response = client.get(
        "/api/v1/metrics/unknown/summary",
        params={"organizational_unit": "Service Nord"},
    )
    upload_response = client.post(
        "/api/v1/imports/csv",
        files={"file": ("metrics.txt", "not csv", "text/plain")},
    )

    assert metric_response.status_code == 404
    assert upload_response.status_code == 415


def test_dashboard_exposes_filters_summaries_and_sources(client: TestClient) -> None:
    client.post("/api/v1/imports/csv", files={"file": ("sample.csv", VALID_CSV, "text/csv")})

    filters = client.get("/api/v1/metrics/dashboard/filters")
    dashboard = client.get(
        "/api/v1/metrics/dashboard",
        params={"organizational_unit": "Service Nord"},
    )

    assert filters.status_code == 200
    assert filters.json() == {
        "organizational_units": ["Service Nord"],
        "period_start": "2026-07-01",
        "period_end": "2026-07-14",
    }
    assert dashboard.status_code == 200
    assert dashboard.json()["source_files"] == ["sample.csv"]
    assert len(dashboard.json()["summaries"]) == 2


def test_measurement_trend_and_organization_comparison(client: TestClient) -> None:
    second_unit = VALID_CSV.replace("Service Nord", "Service Süd").replace(",110,100", ",80,100")
    client.post("/api/v1/imports/csv", files={"file": ("north.csv", VALID_CSV, "text/csv")})
    client.post("/api/v1/imports/csv", files={"file": ("south.csv", second_unit, "text/csv")})

    trend = client.get(
        "/api/v1/metrics/handled_cases/measurements",
        params={
            "organizational_unit": "Service Nord",
            "date_from": "2026-07-01",
            "date_to": "2026-07-14",
        },
    )
    comparison = client.get(
        "/api/v1/metrics/handled_cases/comparison",
        params={"date_from": "2026-07-01", "date_to": "2026-07-14"},
    )

    assert trend.status_code == 200
    assert [item["value"] for item in trend.json()] == ["110.0000", "95.0000"]
    assert comparison.status_code == 200
    assert [item["organizational_unit"] for item in comparison.json()["entries"]] == [
        "Service Nord",
        "Service Süd",
    ]
    assert [item["value"] for item in comparison.json()["entries"]] == ["205.00", "175.00"]


def test_comparison_rejects_unknown_metric(client: TestClient) -> None:
    response = client.get("/api/v1/metrics/unknown/comparison")
    assert response.status_code == 404


def test_xlsx_import_uses_the_same_validated_pipeline(client: TestClient) -> None:
    workbook = Workbook()
    sheet = workbook.active
    for row in VALID_CSV.strip().splitlines():
        sheet.append(row.split(","))
    content = BytesIO()
    workbook.save(content)

    response = client.post(
        "/api/v1/imports/file",
        files={
            "file": (
                "sample.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["status"] == "completed"
    assert response.json()["imported_rows"] == 3


def test_xlsx_import_accepts_original_wide_employee_report(client: TestClient) -> None:
    workbook = Workbook()
    sheet = workbook.active
    headers = [
        "Teamleiter",
        "Mitarbeiter",
        "EPA",
        "VVL",
        "BNT",
        "Mobile",
        "VVL Mobile",
        "Angebote",
        "Angebotsquote",
        "Calls",
        "BBCR",
        "Bewertungen",
        "TNPS",
        "CS",
        "Total Fix",
        "Auflegerquote",
        "CHT",
        "ACW",
        "FB Quote",
    ]
    for column, header in enumerate(headers, start=2):
        sheet.cell(11, column, header)
    values = [
        "TL Nord",
        "Erika Muster",
        "4711",
        2,
        1,
        0,
        1,
        5,
        0.5,
        20,
        0.25,
        8,
        42,
        90,
        0.8,
        0.1,
        365,
        64,
        0.75,
    ]
    for column, value in enumerate(values, start=2):
        sheet.cell(12, column, value)
    for column in (10, 12, 16, 17, 20):
        sheet.cell(12, column).number_format = "0.00%"
    content = BytesIO()
    workbook.save(content)

    response = client.post(
        "/api/v1/imports/file",
        files={
            "file": (
                "employee-report.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["status"] == "completed"
    assert response.json()["imported_rows"] == 16
    filters = client.get("/api/v1/metrics/dashboard/filters").json()
    assert filters["organizational_units"] == ["TL Nord · Erika Muster"]
    offer_rate = client.get(
        "/api/v1/metrics/offer_rate/summary",
        params={"organizational_unit": "TL Nord · Erika Muster"},
    )
    assert offer_rate.json()["value"] == "50.00"


def test_xlsx_import_falls_back_to_wide_daily_summary(client: TestClient) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["E2"] = "AHT"
    sheet["F2"] = "ACW"
    sheet["D3"] = "April"
    sheet["E3"] = 365
    sheet["F3"] = 64
    summary_headers = ["EPA", "VVL", "BNT", "Mobile", "Angebote", "Calls", "BBCR"]
    for column, header in enumerate(summary_headers, start=4):
        sheet.cell(9, column, header)
    for column, value in enumerate(["Potsdam", 2, 1, 0, 4, 20, 0.25], start=4):
        sheet.cell(10, column, value)
    sheet.cell(10, 10).number_format = "0.00%"
    for column, header in enumerate(["Teamleiter", "Mitarbeiter", *summary_headers], start=2):
        sheet.cell(11, column, header)
    content = BytesIO()
    workbook.save(content)

    response = client.post(
        "/api/v1/imports/file",
        files={
            "file": (
                "summary-report.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["status"] == "completed"
    assert response.json()["imported_rows"] == 8
    assert client.get("/api/v1/metrics/dashboard/filters").json()["organizational_units"] == [
        "Potsdam"
    ]
    aht = client.get(
        "/api/v1/metrics/aht/summary",
        params={"organizational_unit": "Potsdam"},
    )
    assert aht.json()["value"] == "365.00"
