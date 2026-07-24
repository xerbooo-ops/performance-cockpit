from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

from performance_cockpit import web


def test_python_dashboard_renders_without_frontend_build(client: TestClient) -> None:
    response = client.get("/")

    assert response.status_code == 200
    assert "Performance Cockpit" in response.text
    assert "Python-only · EPA-Anonymisierung" not in response.text
    assert "Ohne Node.js, TypeScript oder externe Dienste." not in response.text
    assert "● vollständig lokal" not in response.text
    assert 'action="/web/import"' not in response.text
    assert "Datei importieren" not in response.text
    assert 'http-equiv="refresh"' in response.text
    assert "Reportdatei auswählen" in response.text
    assert "alle 5 Sekunden geprüft" in response.text


def test_python_dashboard_imports_and_filters_by_epa(client: TestClient) -> None:
    workbook = Workbook()
    sheet = workbook.active
    headers = ["Teamleiter", "Mitarbeiter", "EPA", "VVL", "BNT", "Calls"]
    for column, header in enumerate(headers, start=2):
        sheet.cell(11, column, header)
    for column, value in enumerate(["Leitung", "Name", "EPA-7", 2, 1, 5], start=2):
        sheet.cell(12, column, value)
    payload = BytesIO()
    workbook.save(payload)

    imported = client.post(
        "/web/import",
        files={
            "file": (
                "report.xlsx",
                payload.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        follow_redirects=True,
    )

    assert imported.status_code == 200
    assert "6 Zeilen importiert" in imported.text
    assert "EPA-7" in imported.text
    assert "Potsdam" in imported.text
    assert "Leitung" not in imported.text
    assert ">Name<" not in imported.text


def test_python_dashboard_reset_requires_delete(client: TestClient) -> None:
    rejected = client.post(
        "/web/reset",
        data={"confirmation": "delete"},
        follow_redirects=True,
    )
    accepted = client.post(
        "/web/reset",
        data={"confirmation": "DELETE"},
        follow_redirects=True,
    )

    assert "Bestätigung war nicht DELETE" in rejected.text
    assert "Alle lokalen Daten wurden gelöscht" in accepted.text


def test_python_dashboard_selects_a_file_for_automatic_updates(
    client: TestClient,
    monkeypatch,
    tmp_path,
) -> None:
    report_path = tmp_path / "report.xlsx"
    report_path.touch()
    selected_paths = []
    monkeypatch.setattr(web, "select_report_file", lambda: report_path)
    monkeypatch.setattr(
        client.app.state.file_watcher,
        "select_file",
        lambda path: selected_paths.append(path) or True,
    )

    response = client.post("/web/watch/select", follow_redirects=True)

    assert response.status_code == 200
    assert "report.xlsx wird jetzt automatisch überwacht" in response.text
    assert selected_paths == [report_path]
